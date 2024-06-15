import json
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import random

with open('../mustard_data/data_split_output/mustard_R_dataset_test.json', 'r') as f:
    R_dataset = json.load(f)
    # randomly sample 30 data points
    sampled_R_dataset = dict(random.sample(sorted(R_dataset.items()), 30))

with open('../mustard_data/data_split_output/mustard_AS_dataset_test.json', 'r') as f:
    AS_dataset = json.load(f)
    sampled_AS_dataset = dict(random.sample(sorted(AS_dataset.items()), 30))


with open('../mustard_data/data_split_output/mustard_U_dataset_test.json', 'r') as f:
    U_dataset = json.load(f)
    sampled_U_dataset = dict(random.sample(sorted(U_dataset.items()), 30))

dataset = {**sampled_R_dataset, **sampled_AS_dataset, **sampled_U_dataset}
dataset = dict(random.sample(sorted(dataset.items()), len(dataset)))

with open('../mustard_data/data_raw/mustard_dataset_test.json', 'r') as f:
    dataset_raw = json.load(f)


data_list = []
for key, data in dataset.items():
    data['id'] = key
    data['image'] = f"../mustard_data/data_raw/images/{key}.jpg"
    new_data = {
        'utterance': data['utterance'],
        'image': data['image'],
        'id': data['id'],
        'label': dataset_raw[key]['sarcasm'],
        'speaker': dataset_raw[key]['speaker'],
        'context': ' '.join(dataset_raw[key]['context']),
    }
    print(dataset_raw[key]['sarcasm'])
    data_list.append(new_data)

data = data_list

# Create a workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Extract and insert header
header = ['id', 'text', 'image', 'sarcasm', 'speaker', 'context', 'RUS', 'image_label', 'text_label']
ws.append(header)

# Insert data
for row in data:
    values = [row['id'], row['utterance'], row['image'], row['label'], row['speaker'], row['context'],  '', '', '']
    ws.append(values)

# Adjust column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Insert images
for index, row in enumerate(data, start=2):
    img_path = row['image']
    img = PILImage.open(img_path)
    img.thumbnail((300, 300))  # Resize image if needed
    img.save(img_path)  # Save the resized image

    excel_img = ExcelImage(img_path)
    img_cell_coordinate = get_column_letter(3) + str(index)
    ws.add_image(excel_img, img_cell_coordinate)

    ws.row_dimensions[index].height = 100

# Create a table in the worksheet
tab = Table(displayName="Table1", ref=f"A1:{get_column_letter(len(header))}{len(data) + 1}")
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)

# Save the workbook
wb.save("output.xlsx")

print("Data exported successfully to output.xlsx!")
